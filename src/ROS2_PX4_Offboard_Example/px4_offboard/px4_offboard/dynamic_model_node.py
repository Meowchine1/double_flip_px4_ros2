import rclpy
from rclpy.node import Node
from geometry_msgs.msg import PoseStamped, Vector3, Quaternion
from visualization_msgs.msg import Marker
from geometry_msgs.msg import Point
from std_msgs.msg import Float32MultiArray
from px4_msgs.msg import (VehicleAttitude, VehicleImu, ActuatorOutputs, ActuatorMotors, ActuatorControls,
                          VehicleLocalPosition,SensorCombined,VehicleAngularVelocity, 
                          VehicleAngularAccelerationSetpoint, VehicleMagnetometer, SensorBaro) # TODO CONNECT VehicleMagnetometer SensorBaro
import numpy as np
from rclpy.qos import QoSProfile, ReliabilityPolicy, HistoryPolicy, DurabilityPolicy
from scipy.spatial.transform import Rotation as R
from rclpy.time import Time
from nav_msgs.msg import Odometry  # inner ROS2 SEKF
from filterpy.kalman import ExtendedKalmanFilter
from datetime import datetime
from sensor_msgs.msg import Imu, MagneticField, FluidPressure
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import casadi as ca

#import matplotlib.pyplot as plt
from std_msgs.msg import String
 
from datetime import datetime
from jax import grad,  jit, lax 
import jax
import jax.numpy as jnp
from jax import grad, jacobian, hessian

from quad_flip_msgs.msg import OptimizedTraj

from rclpy.qos import QoSProfile

# ======= CONSTANTS =======
SEA_LEVEL_PRESSURE = 101325.0
EKF_DT = 0.01
# ======= DRONE CONSTRUCT PARAMETERS =======
MASS = 0.82
INERTIA = np.diag([0.045, 0.045, 0.045])
ARM_LEN = 0.15
K_THRUST = 1.48e-6
K_TORQUE = 9.4e-8
MOTOR_TAU = 0.02
MAX_SPEED = 2100.0
DRAG = 0.1
MAX_RATE = 25.0  # рад/с, ограничение на угловую скорость (roll/pitch)
# ===== MATRIX OPERTIONS =====
# QUATERNION UTILS (SCIPY-based)
def quat_multiply(q1, q2):
    """
    Умножение кватернионов q1 * q2
    q = [w, x, y, z]
    """
    w1, x1, y1, z1 = q1
    w2, x2, y2, z2 = q2
    w = w1*w2 - x1*x2 - y1*y2 - z1*z2
    x = w1*x2 + x1*w2 + y1*z2 - z1*y2
    y = w1*y2 - x1*z2 + y1*w2 + z1*x2
    z = w1*z2 + x1*y2 - y1*x2 + z1*w2
    return jnp.array([w, x, y, z])
def f(x, u, dt):
    m = MASS
    I = INERTIA
    arm = ARM_LEN
    kf = K_THRUST
    km = K_TORQUE
    drag = DRAG  # единственный коэффициент сопротивления
    g = jnp.array([0.0, 0.0, 9.81])
    max_speed = MAX_SPEED

    pos = x[0:3]
    vel = x[3:6]
    quat = x[6:10]  # Кватернионы
    omega = x[10:13]

    # Нормализация кватерниона (проверка на нулевую норму)
    quat_norm = jnp.linalg.norm(quat)
    if quat_norm < 1e-8:
        quat = jnp.array([1.0, 0.0, 0.0, 0.0])  # Если нулевая норма, ставим единичный кватернион
    else:
        quat = quat / quat_norm

    R_bw = jnp.array(R.from_quat(quat).as_matrix())  # матрица поворота из кватерниона

    rpm = jnp.clip(u, 0.0, max_speed)
    w_squared = rpm ** 2
    thrusts = kf * w_squared

    Fz_body = jnp.array([0.0, 0.0, jnp.sum(thrusts)])
    F_world = R_bw @ Fz_body - m * g - drag * vel  # линейное сопротивление
    acc = F_world / m

    new_vel = vel + acc * dt
    new_pos = pos + vel * dt + 0.5 * acc * dt ** 2

    tau = jnp.array([
        arm * (thrusts[1] - thrusts[3]),
        arm * (thrusts[2] - thrusts[0]),
        km * (w_squared[0] - w_squared[1] + w_squared[2] - w_squared[3])
    ])

    omega_cross = jnp.cross(omega, I @ omega)
    omega_dot = jnp.linalg.solve(I, tau - omega_cross)
    new_omega = omega + omega_dot * dt

    omega_quat = jnp.concatenate([jnp.array([0.0]), new_omega])  # Создаем кватернион из угловой скорости
    dq = 0.5 * quat_multiply(quat, omega_quat)  # Вычисляем изменение кватерниона
    new_quat = quat + dq * dt
    new_quat /= jnp.linalg.norm(new_quat + 1e-8)  # Нормализуем кватернион

    x_next = jnp.concatenate([new_pos, new_vel, new_quat, new_omega])  # Собираем новое состояние
    return x_next

# def plot_trajectory(x_traj_opt, x_goal=None, title="Optimized Trajectory"):
#     x = x_traj_opt[:, 0]
#     y = x_traj_opt[:, 1]
#     z = x_traj_opt[:, 2]

#     fig = plt.figure()
#     ax = fig.add_subplot(111, projection='3d')
#     ax.plot(x, y, z, label='Predicted trajectory', color='blue')
#     ax.scatter(x[0], y[0], z[0], color='green', label='Start')
#     ax.scatter(x[-1], y[-1], z[-1], color='red', label='End')

#     if x_goal is not None:
#         ax.scatter(x_goal[0], x_goal[1], x_goal[2], color='orange', label='Goal', s=100, marker='X')

#     ax.set_xlabel('X')
#     ax.set_ylabel('Y')
#     ax.set_zlabel('Z')
#     ax.set_title(title)
#     ax.legend()
#     plt.show()

def publish_trajectory_marker(pub, x_traj_opt, frame_id="map"):
    marker = Marker()
    marker.header.frame_id = frame_id
    marker.header.stamp = node.get_clock().now().to_msg()
    marker.ns = "mpc_trajectory"
    marker.id = 0
    marker.type = Marker.LINE_STRIP
    marker.action = Marker.ADD
    marker.scale.x = 0.02
    marker.color.a = 1.0
    marker.color.r = 0.0
    marker.color.g = 0.5
    marker.color.b = 1.0

    for state in x_traj_opt:
        p = Point()
        p.x, p.y, p.z = state[:3]
        marker.points.append(p)

    pub.publish(marker)

    # self.traj_pub = self.create_publisher(Marker, "/mpc/trajectory", 10)
    #publish_trajectory_marker(self.traj_pub, x_traj_opt)
 
 

class MyEKF(ExtendedKalmanFilter):
    def __init__(self, dim_x, dim_z): 
        super().__init__(dim_x, dim_z)
        self.dt = EKF_DT
        self.f = f

    def predict_x(self, u=np.zeros(4)):
        # Custom fx(x, u, dt) function
        return f(x=self.x, u=u, dt=self.dt)
    #predict new state with dynamic physic model

class DynamicModelNode(Node):
    def __init__(self):
        super().__init__('dynamic_model_node')
        
        #self.get_logger().info(f"DynamicModelNode") 

        qos_profile_for_odom = QoSProfile(depth=10)  # стандартный: reliability=RELIABLE, durability=VOLATILE
        qos_profile = QoSProfile(
            reliability=ReliabilityPolicy.BEST_EFFORT,
            durability=DurabilityPolicy.TRANSIENT_LOCAL,
            history=HistoryPolicy.KEEP_LAST,
            depth=10
        )
        # == == == =PUBLISHERS= == == ==
 
        self.datetime = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.log_base = os.path.join("programm_log", self.datetime)

        # Паблишеры для ekf_filter_node
        self.imu_pub = self.create_publisher(Imu, '/imu/data', qos_profile)
        self.mag_pub = self.create_publisher(MagneticField, '/imu/mag', qos_profile)
        self.baro_pub = self.create_publisher(FluidPressure, '/baro', qos_profile)
        self.ekf_state_pub = self.create_publisher(Float32MultiArray, '/ekf/state', qos_profile)

         

        #trajectory visualisation
        self.traj_pub = self.create_publisher(Marker, "/mpc/trajectory", qos_profile)

        self.server_pub = self.create_publisher(String, '/drone/server_msg', qos_profile)

        # == == == =SUBSCRIBERS= == = ==
        self.create_subscription(SensorCombined, '/fmu/out/sensor_combined', self.sensor_combined_callback, qos_profile)
        self.create_subscription(VehicleAngularVelocity, '/fmu/out/vehicle_angular_velocity', self.angular_velocity_callback, qos_profile)
        self.create_subscription(VehicleAttitude, '/fmu/out/vehicle_attitude', self.vehicle_attitude_callback, qos_profile)
        self.create_subscription(VehicleAngularAccelerationSetpoint,
        '/fmu/out/vehicle_angular_acceleration_setpoint', self.vehicle_angular_acceleration_setpoint_callback, qos_profile)
        self.create_subscription(VehicleImu,'/fmu/out/vehicle_imu',self.vehicle_imu_callback, qos_profile)


        self.create_subscription(ActuatorOutputs, '/fmu/out/actuator_outputs', self.actuator_outputs_callback, qos_profile)
        self.create_subscription(ActuatorControls, '/fmu/out/actuator_controls', self.actuator_controls_callback, qos_profile)
        self.create_subscription(ActuatorMotors, '/fmu/out/actuator_motors', self.actuator_motors_callback, qos_profile) 
 
        
        self.create_subscription(VehicleLocalPosition, '/fmu/out/vehicle_local_position', self.vehicle_local_position_callback, qos_profile)
        self.create_subscription(SensorBaro, '/fmu/out/sensor_baro', self.sensor_baro_callback, qos_profile)
        self.create_subscription(VehicleMagnetometer, '/fmu/out/vehicle_magnetometer', self.vehicle_magnetometer_callback, qos_profile)
        
        # ekf_filter_node data
        self.create_subscription(Odometry, '/odometry/filtered', self.odom_callback, qos_profile_for_odom)

        # == == == =DATA USED IN METHODS= == == == 
        self.angularVelocity = np.zeros(3, dtype=np.float32)
        self.angular_acceleration = np.zeros(3, dtype=np.float32)
        self.vehicleImu_velocity_w = np.zeros(3, dtype=np.float32) # в мировых координатах 
        self.sensorCombined_linear_acceleration = np.zeros(3, dtype=np.float32)
        self.position = np.zeros(3, dtype=np.float32) # drone position estimates with IMU localization
        self.motor_inputs = np.zeros(4, dtype=np.float32)  # в радианах
        self.vehicleAttitude_q = np.array([0.0, 0.0, 0.0, 1.0], dtype=np.float32) # quaternion from topic
        self.magnetometer_data = np.zeros(3, dtype=np.float32)
        self.baro_attitude = 0.0
        self.baro_pressure = 0.0
        self.baro_altitude = 0.0
        self.mag_yaw = 0.0
        
        # FOR SITL TESTING  
        self.vehicleLocalPosition_position = np.zeros(3, dtype=np.float32)
        # ekf_filter_node data
        self.odom_callback_position = np.zeros(3, dtype=np.float32)
        self.odom_callback_orientation = np.zeros(4, dtype=np.float32)

        # ======= OTHER TOPIC DATA ======= 
        # :[TOPIC NAME]_[PARAM NAME] OR [TOPIC NAME] IF PARAM = TOPIC NAME
        self.sensorCombined_angular_velocity = np.zeros(3, dtype=np.float32)
        self.angularVelocity_angular_acceleration = np.zeros(3, dtype=np.float32)
        self.baro_temperature = 0.0 # temperature in degrees Celsius

        # ======= EKF =======
        # вектор состояния (13 штук: позиция, скорость, ориентация (4), угловые скорости).
        # наблюдаемые величины (14 штук) позиция, линейная скорость, mag yaw, baro pressure, 
        self.ekf = MyEKF(dim_x=13, dim_z=14)
        self.ekf.x = np.zeros(13)
        self.ekf.x[6] = 1.0  # qw = 1 (единичный кватернион)
        #Covariance matrix
        self.ekf.P *= 0.1
        #Process noise matrix
        self.ekf.Q = np.diag([
            0.001, 0.001, 0.001,      # x, y, z (позиция)
            0.01, 0.01, 0.01,         # vx, vy, vz (скорость)
            0.0001, 0.0001, 0.0001, 0.0001, # qw, qx, qy, qz (ориентация)
            0.00001, 0.00001, 0.00001, # wx, wy, wz (угловая скорость)
            1.25e-7,                  # момент (Nm)^2 s
            0.0005,                   # сила (N)^2 s
        ])
        #Measurement noise matrix
        self.ekf.R = np.diag([
            0.1, 0.1, 0.1,            # позиция x, y, z (м²)
            0.0001, 0.0001, 0.0001,   # скорость vx, vy, vz (м²/с²)
            0.00001, 0.00001, 0.00001, 0.00001, # ориентация qw, qx, qy, qz (кватернионы)
            0.00001, 0.00001, 0.00001, # угловые скорости wx, wy, wz (рад²/с²)
            0.5                        # барометр (м²)
        ])
        #    ====    ====   Параметры ModelPredictiveController    ====     ====     ====
        self.dt = 0.1  # Шаг времени (с)
        # TODO FROM FILE  
        self.horizon = 50  # Горизонт предсказания
        self.n = 13  # Размерность состояния квадрокоптера (позиция, скорость, ориентация, угловая скорость)
        self.m = 4  # Размерность управления (4 мотора)
        
        # стоимости
        # x = [x, y, z,      # позиция
        # vx, vy, vz,   # скорость
        # qw, qx, qy, qz,  # ориентация (кватернион)
        # wx, wy, wz]   # угловая скорость
        self.Q = 100.0  # важность траектории (позиция + ориентация) 
        self.R = 0.001   # разрешаем моторам быть агрессивными
        self.Qf = np.diag([
            10.0, 10.0, 10.0,      # позиция — средне важно
            0.1, 0.1, 0.1,         # скорость — не критично
            0.0, 100.0, 100.0, 0.0, # ориентация — важны qx/qy (например, pitch flip)
            10.0, 10.0, 1.0        # угловые скорости — важно, чтобы стабилизировался
        ])  
        # Инициализация ModelPredictiveController с оптимизатором
        self.mpc = ModelPredictiveController(
            dt=self.dt,
            horizon=self.horizon,
            n=self.n,
            m=self.m,
            Q=self.Q,
            R=self.R,
            Qf=self.Qf
        )

        self.phase = 'init'
        self.takeoff_altitude = 5.0  # м
        self.takeoff_tol = 0.1
        self.flip_started_time = None
        self.flip_duration = 1.0  # с, продолжительность флипа
        self.recovery_time = 2.0  # с, стабилизация после флипа
        self.recovery_start_time = None
        self.landing_altitude = 0.2  # м
        # Переход в посадку — по ориентации
        self.roll_abs_tol = 0.1  # допуск 0.1 рад
        #   ====    ====    ====     ====     ====

        now_str = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        self.log_file_name_pos = f"{now_str}_pos.log"
        self.log_file_name_quat = f"{now_str}_quat.log"
        self.log_file_name_vel = f"{now_str}_vel.log"
        self.log_file_name_ang_vel = f"{now_str}_ang_vel.log"

         # ======= TIMERS =======
        #self.timer = self.create_timer(0.01, self.step_dynamics)
        self.EKF_timer = self.create_timer(EKF_DT, self.EKF)
        self.ekf_filter_node_timer = self.create_timer(0.01, self.ekf_filter_node_t) 
        self.mpc_controller = self.create_timer(0.01, self.mpc_control_loop)
         
        # == == == =CLIENT SERVER INTERACTION= == == == 
        self.create_subscription(String, '/drone/client_msg', self.client_msg_callback, qos_profile)#
        self.pub_optimized_traj = self.create_publisher(OptimizedTraj, '/drone/optimized_traj', qos_profile)


        #self.optimized_traj = self.create_timer(EKF_DT, self.send_optimized_traj_t)
        self.optimized_traj_f = False
        self.X_opt = np.zeros((self.horizon + 1, self.n))  # (N+1) x n
        self.u_optimal = np.zeros((self.horizon, self.m))  # N x m
        self.i_final = 0
        self.cost_final = 0.0
        self.done = False 
        self.to_client_f = False


    # =============================нет колбека========================================
    def actuator_outputs_callback(self, msg: ActuatorOutputs):
        pwm_outputs = msg.output[:4]  # предполагаем, что 0-3 — это моторы
        # Преобразование PWM в радианы в секунду (линейное приближение)
         
        self.motor_inputs = np.clip((np.array(pwm_outputs) - 1000.0) / 1000.0 * MAX_SPEED, 0.0, MAX_SPEED)
        self.get_logger().info(f"actuator_outputs_callback {msg}")

    
 
    def actuator_motors_callback(self, msg):
        self.get_logger().info(f"actuator_outputs_callback {msg}")


    def client_msg_callback(self, msg):#
        """GET CLIENT MESSAGES"""
        command = msg.data.strip().lower()
        self.get_logger().info(f"Received command: {command}")
        if command == "takeoff":
            self.phase = command
            self.to_client_f = True
            self.optimized_traj_f = True 
        else:
            self.get_logger().warn(f"Unknown command: {command}") 


    def odom_callback(self, msg: Odometry):
        self.get_logger().info("odom_callback")
        self.odom_callback_position = msg.pose.pose.position
        self.odom_callback_orientation = msg.pose.pose.orientation


    # ==============================нет колбека=======================================

    def send_optimized_traj(self):
        if self.optimized_traj_f:
            msg = OptimizedTraj()
            msg.x_opt = np.asarray(self.X_opt).flatten().astype(np.float32).tolist()
            msg.u_opt = np.asarray(self.u_optimal).flatten().astype(np.float32).tolist()
            msg.i_final = int(self.i_final)
            msg.cost_final = float(self.cost_final)
            msg.done = self.done
            self.pub_optimized_traj.publish(msg)

 
 
    # def publish_trajectory_marker(pub, x_traj_opt, frame_id="map"):
    #     marker = Marker()
    #     marker.header.frame_id = frame_id
    #     marker.header.stamp = self.get_clock().now().to_msg()#?
    #     marker.ns = "mpc_trajectory"
    #     marker.id = 0
    #     marker.type = Marker.LINE_STRIP
    #     marker.action = Marker.ADD
    #     marker.scale.x = 0.02
    #     marker.color.a = 1.0
    #     marker.color.r = 0.0
    #     marker.color.g = 0.5
    #     marker.color.b = 1.0

    #     for state in x_traj_opt:
    #         p = Point()
    #         p.x, p.y, p.z = state[:3]
    #         marker.points.append(p)

    #     pub.publish(marker)
     
    def quaternion_from_roll(self, roll_rad):
        r = R.from_euler('x', roll_rad)
        return r.as_quat()  # формат [x, y, z, w]
    
    def send_msg_to_client(self, msg):
        server_msg = String()
        server_msg.data = msg 
        self.server_pub.publish(server_msg)

    def mpc_control_loop(self):
        if self.phase != 'init':
            current_time = self.get_clock().now().nanoseconds * 1e-9
            x0 = self.ekf.x.copy()  # [13] x, y, z, vel, q, omega
            u_init = jnp.tile(self.motor_inputs, (self.horizon, 1))  # [horizon, 4]
            x_target_traj = jnp.zeros((self.horizon, 13))
            u_target_traj = u_target_traj = jnp.tile(self.motor_inputs, (self.horizon, 1)) #u_target_traj = jnp.zeros((self.horizon, 4))

            if self.phase == 'takeoff':
                for i in range(self.horizon):
                    pos = x0[0:3].copy()
                    pos = pos.at[2].set(self.takeoff_altitude)
                    vel = jnp.zeros(3)
                    q = jnp.array([0.0, 0.0, 0.0, 1.0])
                    omega = jnp.zeros(3)
                    x_target_traj = x_target_traj.at[i].set(jnp.concatenate([pos, vel, q, omega]))
                    u_target_traj = u_target_traj.at[i].set(self.motor_inputs.copy())

                if abs(self.ekf.x[2] - self.takeoff_altitude) < self.takeoff_tol:
                    self.phase = 'flip'
                    self.flip_started_time = current_time
                    self.send_msg_to_client("flip")

            elif self.phase == 'flip':
                t_local = current_time - self.flip_started_time
                t_local = jnp.clip(t_local, 0.0, self.flip_duration)

                roll_expected = 2 * jnp.pi * t_local / self.flip_duration
                q_current = self.ekf.x[6:10]
                roll_current, _, _ = self.euler_from_quaternion(q_current)
                roll_error = roll_expected - roll_current

                gain_base = 0.8
                gain_adaptive = gain_base + 0.3 * jnp.tanh(roll_error)
                roll_target = roll_current + gain_adaptive * roll_error

                for i in range(self.horizon):
                    alpha_i = i / self.horizon
                    angle_i = roll_target * alpha_i

                    pos = x0[0:3]
                    vel = jnp.zeros(3)
                    q = self.quaternion_from_roll(angle_i)

                    omega_magnitude = 2 * jnp.pi / self.flip_duration + 0.2 * roll_error
                    omega = jnp.array([omega_magnitude, 0.0, 0.0])

                    x_target_traj = x_target_traj.at[i].set(jnp.concatenate([pos, vel, q, omega]))
                    u_target_traj = u_target_traj.at[i].set(self.recovery_thrust.copy())

                if abs(roll_current) >= 2 * jnp.pi * 0.95:
                    self.phase = 'recovery'
                    self.recovery_start_time = current_time

            elif self.phase == 'recovery':
                t_local = current_time - self.recovery_start_time
                t_local = jnp.clip(t_local, 0.0, self.recovery_time)

                roll_desired = 2 * jnp.pi * (1 - t_local / self.recovery_time)
                q_current = self.ekf.x[6:10]
                roll_current, _, _ = self.euler_from_quaternion(q_current)
                roll_error = roll_desired - roll_current

                gain = 0.6 + 0.4 * (abs(roll_error) / jnp.pi)
                roll_target = roll_current + gain * roll_error

                for i in range(self.horizon):
                    alpha_i = i / self.horizon
                    angle_i = roll_current + alpha_i * (roll_target - roll_current)

                    pos = x0[0:3].copy()
                    vel = jnp.zeros(3)
                    q = self.quaternion_from_roll(angle_i)

                    omega_mag = -2 * jnp.pi / self.recovery_time * (1 + 0.2 * abs(roll_error) / jnp.pi)
                    omega = jnp.array([omega_mag, 0.0, 0.0])

                    x_target_traj = x_target_traj.at[i].set(jnp.concatenate([pos, vel, q, omega]))
                    u_target_traj = u_target_traj.at[i].set(self.recovery_thrust.copy())

                if abs(roll_current) <= self.roll_abs_tol:
                    self.phase = 'land'

            elif self.phase == 'land':
                self.to_client_f = False
                self.optimized_traj_f = False
                self.done = True
                self.send_msg_to_client("land")

            # Вызов MPC
            self.X_opt, self.u_optimal, self.i_final, self.cost_final = self.mpc.step(
                #t=current_time,
                x0=x0,
                u_init=u_init,
                x_target_traj=x_target_traj[-1],
                u_target_traj=u_target_traj,
            )
            self.send_optimized_traj()
            
            #publish_trajectory_marker(self.traj_pub, self.X_opt)

    def ekf_filter_node_t(self):
        #self.get_logger().info("ekf_filter_node_t")
        
        def to_float_array(arr):
            return [float(x) for x in arr]

        imu_msg = Imu()
        mag_msg = MagneticField()
        baro_msg = FluidPressure()

        # Стамп времени для сообщений
        current_time = self.get_clock().now().to_msg()

        # ======== /imu/data ========
        imu_msg.header.stamp = current_time
        imu_msg.header.frame_id = "base_link"

        # Ориентация
        q = to_float_array(self.vehicleAttitude_q)
        imu_msg.orientation = Quaternion(x=q[0], y=q[1], z=q[2], w=q[3])

        # Угловая скорость
        ang_vel = to_float_array(self.angularVelocity)
        imu_msg.angular_velocity = Vector3(x=ang_vel[0], y=ang_vel[1], z=ang_vel[2])

        # Линейное ускорение
        lin_acc = to_float_array(self.sensorCombined_linear_acceleration)
        imu_msg.linear_acceleration = Vector3(x=lin_acc[0], y=lin_acc[1], z=lin_acc[2])

        self.imu_pub.publish(imu_msg)

        # ======== /imu/mag ========
        mag_msg.header.stamp = current_time
        mag_msg.header.frame_id = "base_link"
        mag = to_float_array(self.magnetometer_data)
        mag_msg.magnetic_field = Vector3(x=mag[0], y=mag[1], z=mag[2])
        self.mag_pub.publish(mag_msg)

        # ======== /baro ========
        baro_msg.header.stamp = current_time
        baro_msg.header.frame_id = "base_link"
        baro_msg.fluid_pressure = float(self.baro_pressure)
        self.baro_pub.publish(baro_msg)

    # def publish_ekf_state(self):
    #     # Создание сообщения
    #     msg = Float32MultiArray()
    #     msg.data = self.ekf.x.tolist()  # Преобразуем numpy массив в список для публикации

    #     # Публикуем сообщение
    #     self.ekf_state_pub.publish(msg)
    #     #self.get_logger().info("Published EKF state")

    def ekf_logger(self):

        pos_my_ekf = self.ekf.x[0:3]
        pos_odom = self.odom_callback_position
        pos_real = self.vehicleLocalPosition_position

        quat_my_ekf = self.ekf.x[6:10]
        px4_quat = self.vehicleAttitude_q
        quat_odom = self.odom_callback_orientation

        vel_my_ekf = self.ekf.x[3:6]
        integral_vel = self.vehicleImu_velocity_w

        omega_my_ekf = self.ekf.x[10:13]
        omega_from_sensor = self.angularVelocity

        log_base = self.log_base  # уже должен содержать папку с timestamp

        self._write_to_excel(
            os.path.join(log_base, 'pos_log.xlsx'),
            ['pos_my_ekf', 'pos_odom', 'pos_real'],
            [pos_my_ekf, pos_odom, pos_real],
            error_pairs=[(0, 2), (1, 2)]
        )

        self._write_to_excel(
            os.path.join(log_base, 'quat_log.xlsx'),
            ['quat_my_ekf', 'px4_quat', 'quat_odom'],
            [quat_my_ekf, px4_quat, quat_odom],
            error_pairs=[(0, 1), (2, 1)]
        )

        self._write_to_excel(
            os.path.join(log_base, 'vel_log.xlsx'),
            ['vel_my_ekf', 'integral_vel'],
            [vel_my_ekf, integral_vel],
            error_pairs=[(0, 1)]
        )

        self._write_to_excel(
            os.path.join(log_base, 'ang_vel_log.xlsx'),
            ['omega_my_ekf', 'omega_from_sensor'],
            [omega_my_ekf, omega_from_sensor],
            error_pairs=[(0, 1)]
        )

    def _write_to_excel(self, file_path, labels, data, error_pairs=None):
        """
        Записывает данные в Excel-файл, добавляя колонки ошибок между заданными парами векторов.
        error_pairs: список пар индексов, между которыми считать разность (vec[i] - vec[j])
        """
        os.makedirs(os.path.dirname(file_path), exist_ok=True)

        if error_pairs is None:
            error_pairs = []

        new_file = not os.path.exists(file_path)

        if new_file:
            wb = Workbook()
            ws = wb.active
            ws.title = "Log"

            headers = []
            for i, (label, arr) in enumerate(zip(labels, data)):
                if len(arr) > 1:
                    headers.extend([f"{label}[{j}]" for j in range(len(arr))])
                else:
                    headers.append(label)

            for i, j in error_pairs:
                label_i, label_j = labels[i], labels[j]
                arr_len = len(data[i])
                headers.extend([f"{label_i}-{label_j}[{k}]" for k in range(arr_len)])

            ws.append(headers)
            wb.save(file_path)

        wb = load_workbook(file_path)
        ws = wb.active

        row_values = [float(v) for arr in data for v in arr]

        for i, j in error_pairs:
            diff = np.array(data[i]) - np.array(data[j])
            row_values.extend([float(v) for v in diff])

        ws.append(row_values)
        wb.save(file_path)
 
    def sensor_baro_callback(self, msg):
        #self.get_logger().info("sensor_baro_callback")
        self.baro_temperature = msg.temperature
        self.baro_pressure = msg.pressure
        self.baro_attitude = 44330.0 * (1.0 - (msg.pressure / SEA_LEVEL_PRESSURE) ** 0.1903)

    def get_yaw_from_mag(self):
        r = R.from_quat(self.vehicleAttitude_q)  # [x, y, z, w]
    
        # Преобразуем магнитное поле в мировую систему координат
        mag_world = r.apply(self.magnetometer_data)

        # Проекция магнитного поля на горизонтальную плоскость (X, Y)
        mag_x = mag_world[0]
        mag_y = mag_world[1]
        # Вычисляем yaw (в радианах)
        yaw_from_magnetometer = np.arctan2(-mag_y, mag_x)
        return yaw_from_magnetometer

    def vehicle_magnetometer_callback(self, msg: VehicleMagnetometer):
        # Измерения магнитометра
        #self.get_logger().info("vehicle_magnetometer_callback")
        self.magnetometer_data = np.array(msg.magnetometer_ga, dtype=np.float32)
        self.mag_yaw = self.get_yaw_from_mag()

    # ПОЗИЦИЯ ДЛЯ ОЦЕНКИ ИНЕРЦИАЛНОЙ ЛОКАЛИЗАЦИИ
    def vehicle_local_position_callback(self, msg: VehicleLocalPosition):
        #self.get_logger().info(f"vehicle_local_position_callback {msg.x} {msg.y} {msg.z}")
        self.vehicleLocalPosition_position[0] = msg.x
        self.vehicleLocalPosition_position[1] = msg.y
        self.vehicleLocalPosition_position[2] = msg.z 

    # ЛИНЕЙНОЕ УСКОРЕНИЕ, УГЛОВОЕ УСКОРЕНИЕ, КВАТЕРНИОН
    def sensor_combined_callback(self, msg: SensorCombined):
        dt_gyro = msg.gyro_integral_dt * 1e-6  # микросекунды -> секунды
        gyro_rad = np.array(msg.gyro_rad, dtype=np.float32)  # угловая скорость (рад/с)
        self.sensorCombined_angular_velocity = gyro_rad
         
        delta_angle = gyro_rad * dt_gyro # Угловое приращение (рад)
        self.sensorCombined_delta_angle = delta_angle
        self.sensorCombined_linear_acceleration = np.array(msg.accelerometer_m_s2, dtype=np.float32)
         
    def angular_velocity_callback(self, msg: VehicleAngularVelocity):
        self.angularVelocity = np.array(msg.xyz, dtype=np.float32)
        self.angularVelocity_angular_acceleration = np.array(msg.xyz_derivative, dtype=np.float32)

    def vehicle_attitude_callback(self, msg: VehicleAttitude):
        # In this system we use scipy format for quaternion. 
        # PX4 topic uses the Hamilton convention, and the order is q(w, x, y, z). So we reorder it
        self.vehicleAttitude_q = np.array([msg.q[1], msg.q[2], msg.q[3], msg.q[0]], dtype=np.float32)
        
    def vehicle_angular_acceleration_setpoint_callback(self, msg: VehicleAngularAccelerationSetpoint):
        self.angular_acceleration = msg.xyz

    def vehicle_imu_callback(self, msg: VehicleImu):
        delta_velocity = np.array(msg.delta_velocity, dtype=np.float32)  # м/с
        delta_velocity_dt = msg.delta_velocity_dt * 1e-6  # с
        # Проверяем наличие ориентации и валидного времени интеграции
        if delta_velocity_dt > 0.0:
            rotation = R.from_quat(self.vehicleAttitude_q)
            delta_velocity_world = rotation.apply(delta_velocity)
            gravity = np.array([0.0, 0.0, 9.80665], dtype=np.float32)
            delta_velocity_world += gravity * delta_velocity_dt
            self.vehicleImu_velocity_w += delta_velocity_world
            self.position += self.vehicleImu_velocity_w * delta_velocity_dt

    def publish_motor_inputs(self):
        msg = Float32MultiArray()
        msg.data = self.motor_inputs.tolist()
        self.motor_pub.publish(msg)

    def EKF(self):
        #self.get_logger().info("EKF")
        """ Основная функция обновления фильтра Калмана. """
        vel_world = self.vehicleImu_velocity_w
        z = np.array([
            self.position[0],  # x
            self.position[1],  # y
            self.position[2],  # z
            vel_world[0],  # vx
            vel_world[1],  # vy
            vel_world[2],  # vz
            self.vehicleAttitude_q[0],
            self.vehicleAttitude_q[1],
            self.vehicleAttitude_q[2],
            self.vehicleAttitude_q[3],
            self.angularVelocity[0],
            self.angularVelocity[1],
            self.angularVelocity[2],
            self.baro_altitude  # высота по барометру
        ])

        # Вызываем предсказание и обновление фильтра
        #self.ekf.predict_update(z, HJacobian=self.HJacobian, Hx=self.hx, u=self.motor_inputs)
        self.ekf.x = self.ekf.predict_x(self.motor_inputs)
        self.ekf.update(z, HJacobian=self.HJacobian, Hx=self.hx)
        self.ekf_logger()

    def hx(self, x):
        """ Модель измерений: что бы показали датчики при текущем состоянии. """
        return np.array([
        x[0],  # x (позиция)
        x[1],  # y (позиция)
        x[2],  # z (позиция)
        x[3],  # vx (скорость)
        x[4],  # vy (скорость)
        x[5],  # vz (скорость)
        x[6],  # qw (ориентация)
        x[7],  # qx (ориентация)
        x[8],  # qy (ориентация)
        x[9],  # qz (ориентация)
        x[10], # wx (угловая скорость)
        x[11], # wy (угловая скорость)
        x[12], # wz (угловая скорость)
        x[2],  # барометр (ещё раз высота z)
    ])

    def HJacobian(self, x):
        """ Якобиан модели измерений по состоянию. """
        H = np.zeros((14, 13))  # 14 измерений на 13 состояний
        H[0, 0] = 1.0  # x
        H[1, 1] = 1.0  # y
        H[2, 2] = 1.0  # z
        H[3, 3] = 1.0  # vx
        H[4, 4] = 1.0  # vy
        H[5, 5] = 1.0  # vz
        H[6, 6] = 1.0  # qw
        H[7, 7] = 1.0  # qx
        H[8, 8] = 1.0  # qy
        H[9, 9] = 1.0  # qz
        H[10, 10] = 1.0  # wx
        H[11, 11] = 1.0  # wy
        H[12, 12] = 1.0  # wz
        H[13, 2] = 1.0   # барометрический z
        return H

# # Функция для предсказания успешности флипа
# def predict_flip_success(x_init, u_init, duration=2.0):
#     """
#     Предсказание успешности флипа.
    
#     x_init: Начальное состояние (позиция, скорость, ориентация, угловая скорость).
#     u_init: Управление (обороты моторов).
#     duration: Длительность флипа в секундах.
    
#     Возвращает True, если флип успешен, False - если нет.
#     """
#     # Инициализация переменных
#     x = np.copy(x_init)
#     t = 0
#     time_steps = int(duration / DT)
    
#     # Симуляция траектории на несколько шагов вперед
#     for step in range(time_steps):
#         x = dynamic_model(x, u_init, DT)  # Обновляем состояние по динамической модели
        
#         # Проверяем ограничения
#         pos = x[0:3]
#         vel = x[3:6]
#         quat = x[6:10]
#         omega = x[10:13]
        
#         # Ограничение по высоте
#         if pos[2] < 0 or pos[2] > MAX_HEIGHT:
#             return False  # Если высота выходит за пределы, флип неуспешен
        
#         # Ограничение по угловой скорости
#         if np.abs(omega[0]) > MAX_RATE or np.abs(omega[1]) > MAX_RATE or np.abs(omega[2]) > MAX_RATE:
#             return False  # Если угловая скорость превышает допустимую, флип неуспешен
        
#         # Ограничение по вертикальной скорости
#         if np.abs(vel[2]) > 10:  # можно изменить это значение, если нужно
#             return False  # Если вертикальная скорость слишком велика, флип неуспешен
    
#     # Если все ограничения соблюдены, флип успешен
#     return True 


class ILQROptimizer:
    def __init__(self, dynamics_func, cost_function_traj_flip, horizon, n, m, dt):
        self.f = dynamics_func
        self.cost_function_traj_flip = cost_function_traj_flip
        self.horizon = horizon
        self.n = n
        self.m = m
        self.dt = dt

    def simulate_trajectory(self, x0, U):
        X = [x0]
        x = x0
        for u in U:
            x = self.f(x, u, self.dt)
            X.append(x)
        return jnp.stack(X)

    def linearize_dynamics(self, x, u):
        A = jacobian(self.f, argnums=0)(x, u, self.dt)
        B = jacobian(self.f, argnums=1)(x, u, self.dt)
        return A, B

    def quadratize_cost(self, x, u, x_target, u_target, Q, R):
        lx = grad(lambda x_: self.cost_function_traj_flip(x_[None, :], u[None, :], x_target[None, :], u_target[None, :], Q, R, jnp.zeros((self.n, self.n))))(x)
        lu = grad(lambda u_: self.cost_function_traj_flip(x[None, :], u_[None, :], x_target[None, :], u_target[None, :], Q, R, jnp.zeros((self.n, self.n))))(u)
        lxx = hessian(lambda x_: self.cost_function_traj_flip(x_[None, :], u[None, :], x_target[None, :], u_target[None, :], Q, R, jnp.zeros((self.n, self.n))))(x)
        luu = hessian(lambda u_: self.cost_function_traj_flip(x[None, :], u_[None, :], x_target[None, :], u_target[None, :], Q, R, jnp.zeros((self.n, self.n))))(u)
        lux = jacobian(lambda u_: grad(lambda x_: self.cost_function_traj_flip(x_[None, :], u_[None, :], x_target[None, :], u_target[None, :], Q, R, jnp.zeros((self.n, self.n))))(x))(u)
        return lx, lu, lxx, luu, lux

    def backward_pass(self, X, U, x_target_traj, u_target_traj, Q, R, Qf):
        Vx = grad(lambda x: jnp.dot((x - x_target_traj[-1]), Qf @ (x - x_target_traj[-1])))(X[-1])
        Vxx = Qf
        K_list = []
        k_list = []

        for k in reversed(range(self.horizon)):
            xk = X[k]
            uk = U[k]
            xt = x_target_traj[k]
            ut = u_target_traj[k]

            A, B = self.linearize_dynamics(xk, uk)
            lx, lu, lxx, luu, lux = self.quadratize_cost(xk, uk, xt, ut, Q, R)

            Qx = lx + A.T @ Vx
            Qu = lu + B.T @ Vx
            Qxx = lxx + A.T @ Vxx @ A
            Quu = luu + B.T @ Vxx @ B
            Qux = lux + B.T @ Vxx @ A

            Quu_reg = Quu + 1e-6 * jnp.eye(self.m)  # регуляризация
            Quu_inv = jnp.linalg.inv(Quu_reg)

            K = -Quu_inv @ Qux
            kff = -Quu_inv @ Qu

            Vx = Qx + K.T @ Quu @ kff + K.T @ Qu + Qux.T @ kff
            Vxx = Qxx + K.T @ Quu @ K + K.T @ Qux + Qux.T @ K

            K_list.insert(0, K)
            k_list.insert(0, kff)

        return K_list, k_list

    def forward_pass(self, X, U, k_list, K_list, alpha):
        x = X[0]
        X_new = [x]
        U_new = []
        for k in range(self.horizon):
            dx = x - X[k]
            du = k_list[k] + K_list[k] @ dx
            u_new = U[k] + alpha * du
            U_new.append(u_new)
            x = self.f(x, u_new, self.dt)
            X_new.append(x)
        return jnp.stack(X_new), jnp.stack(U_new)

    def solve(self, x0, u_init, x_target_traj, u_target_traj, Q, R, Qf,
              max_iters=100, tol=1e-3, alpha=1.0):
        X = self.simulate_trajectory(x0, u_init)
        U = u_init

        for i in range(max_iters):
            cost_prev = self.cost_function_traj_flip(X, U, x_target_traj, u_target_traj, Q, R, Qf)
            K_list, k_list = self.backward_pass(X, U, x_target_traj, u_target_traj, Q, R, Qf)
            X_new, U_new = self.forward_pass(X, U, k_list, K_list, alpha)
            cost_new = self.cost_function_traj_flip(X_new, U_new, x_target_traj, u_target_traj, Q, R, Qf)

            if jnp.abs(cost_prev - cost_new) < tol:
                break
            X, U = X_new, U_new

        return X, U, i, cost_new


class ModelPredictiveController:
    def __init__(self, dt, horizon, Q, R, Qf, n, m):
        """
        f: функция динамики: f(x, u, dt) → x_next
        fx_batch: функция для вычисления Якобианов A, B по заданному состоянию и управлению
        dt: шаг по времени
        horizon: длина горизонта предсказания
        Q, R, Qf: матрицы весов стоимости
        n: размерность состояния
        m: размерность управления
        """
        self.dt = dt
        self.horizon = horizon
        self.Q = Q
        self.R = R
        self.Qf = Qf
        self.n = n
        self.m = m
        # Инициализация ILQR оптимизатора
        self.optimizer = ILQROptimizer(
            dynamics_func=f, 
            cost_function_traj_flip=self.cost_function_traj_flip, 
            horizon=horizon, 
            n=n, 
            m=m, 
            dt=dt)
          
    def cost_function_traj_flip(x_traj, u_traj, x_target_traj, u_target_traj, Q, R, Qf):
        def compute_step_cost(x, u, x_target, u_target):
            position_error = x[0:3] - x_target[0:3]
            q_current = x[6:10] / jnp.linalg.norm(x[6:10])
            q_target = x_target[6:10] / jnp.linalg.norm(x_target[6:10])
            dot_product = jnp.clip(jnp.dot(q_current, q_target), -1.0, 1.0)
            orientation_error = 2.0 * jnp.arccos(jnp.abs(dot_product))
            control_error = u - u_target
            return (position_error @ position_error) * Q + orientation_error**2 * Q + (control_error @ control_error) * R

        # Суммарная промежуточная стоимость
        total_cost = jnp.sum(jax.vmap(compute_step_cost)(
            x_traj[:-1], u_traj, x_target_traj[:-1], u_target_traj))

        # Стоимость терминального состояния
        terminal_error = x_traj[-1] - x_target_traj[-1]
        terminal_cost = terminal_error @ Qf @ terminal_error

        return total_cost + terminal_cost
    
    def step(self, x0, u_init, x_target_traj, u_target_traj):
        """
        Вычисляет оптимальную траекторию состояний и управляющих воздействий
        от текущего состояния x0, используя iLQR.
        """ 
        # Используем ILQR для расчета оптимальной траектории
        X_opt, U_opt, i_final, cost_final = self.optimizer.solve( 
            x0=x0,
            u_init=u_init,
            Q=self.Q,
            R=self.R,
            Qf=self.Qf,
            x_target_traj=x_target_traj,
            u_target_traj=u_target_traj
        ) 
        # MPC работает по принципу "перепланирования" на каждом шаге
        # Выбираем управляющее воздействие на текущем шаге
        u_optimal = U_opt[0]  # Выбираем первое управление из оптимизированной траектории

        # Возвращаем оптимальные состояния, управляющие воздействия и информацию о стоимости
        return X_opt, u_optimal, i_final, cost_final

   
 
def main(args=None):
    rclpy.init(args=args)
    node = DynamicModelNode()
    rclpy.spin(node)
    node.destroy_node()
    rclpy.shutdown()

if __name__ == '__main__':
    main()

